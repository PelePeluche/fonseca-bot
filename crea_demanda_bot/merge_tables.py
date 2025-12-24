import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re  # Librería para expresiones regulares

# Cargar los datos de las tablas
tabla_1 = pd.read_excel("TITULOS TF 2024 - SONZINI con Orden y Año.xlsx", sheet_name="Hoja1")
tabla_2 = pd.read_excel("nuevos_datos_extraidos_completos.xlsx", sheet_name="Sheet1")

# Renombrar columnas para facilitar el merge
tabla_1.rename(columns={"CAUSA": "Causa Nro.", "ID": "Identificador", "CARATULA": "Nombre o Razón Social"}, inplace=True)

# Realizar el merge
merged_table = pd.merge(tabla_1, tabla_2, on="Causa Nro.", how="outer")

# Corregir la columna Domicilio
def clean_domicilio(value):
    if isinstance(value, str):
        # Reemplazar N seguido de un carácter incorrecto (*, ", ?) por N°
        return re.sub(r'N[*”?:]', 'N°', value)
    return value

merged_table["Domicilio"] = merged_table["Domicilio"].apply(clean_domicilio)

# Crear las columnas finales
merged_table["Identificador"] = merged_table["Identificador_x"].combine_first(merged_table["Identificador_y"])
merged_table["Nombre o Razón Social"] = merged_table["Nombre o Razón Social_x"]
merged_table["CUIT"] = merged_table["CUIT_x"]
merged_table["PROCURADOR"] = merged_table["PROCURADOR"]
merged_table["Repartición"] = "Tribunal de falta"

# Procesar la columna ORDEN/AÑO para separar en Orden y Año
def split_order_year(value):
    if isinstance(value, str) and '/' in value:
        parts = value.split('/')
        if len(parts) == 3:
            return parts[1], parts[2]
    return None, None

merged_table[["Orden", "Año"]] = merged_table["ORDEN/AÑO"].apply(lambda x: pd.Series(split_order_year(x)))

# Seleccionar las columnas finales incluyendo las requeridas de tabla 2
columns_to_keep = [
    "Identificador", "Nombre o Razón Social", "Causa Nro.", "CUIT", "Orden", "Año", "PROCURADOR",
    "Archivo", "Domicilio", "Matrícula", "Monto total", "Tipo de Persona",
    "Juzgado Nro.", "Juez", "Secretaria", "Fecha", "Fecha de Última Notificación", "Repartición"
]
final_table = merged_table[columns_to_keep]

# Guardar la tabla procesada en un archivo Excel
output_path = "Merged_Table_Processed_Cleaned.xlsx"
final_table.to_excel(output_path, index=False)

# Aplicar formato con openpyxl
wb = load_workbook(output_path)
ws = wb.active

# Estilos
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
border_style = Side(style="thin", color="000000")

# Formato de encabezados
for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

# Formato de celdas
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

# Ajustar el ancho de las columnas automáticamente
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Obtener la letra de la columna
    for cell in column:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

# Guardar el archivo con formato
wb.save(output_path)

print("Tabla procesada, limpiada y guardada con formato como 'Merged_Table_Processed_Cleaned.xlsx'")
