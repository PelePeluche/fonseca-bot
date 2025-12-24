import os
import re
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def read_pdf_with_ocr(pdf_path):
    """Extracts text from a PDF file using OCR."""
    all_text = ""
    try:
        images = convert_from_path(pdf_path)
        for image in images:
            text = pytesseract.image_to_string(image, lang='spa')
            all_text += text + "\n"
    except Exception as e:
        print(f"Error al procesar OCR para {pdf_path}: {e}")
    return all_text

def extract_fields_from_text(text, filename):
    """Extracts relevant fields from the extracted text."""
    fields = {'Archivo': filename}
    try:
        fields['Nombre o Razón Social'] = re.search(r'Nombre\s*:\s*(.+)', text).group(1).strip() if re.search(r'Nombre\s*:\s*(.+)', text) else None
        fields['CUIT'] = re.search(r'NroDoc\.\s*:\s*(\d+)', text).group(1) if re.search(r'NroDoc\.\s*:\s*(\d+)', text) else None
        fields['Repartición'] = None
        fields['Orden'] = None
        fields['Año'] = re.search(r'Año\s*:\s*(\d+)', text).group(1) if re.search(r'Año\s*:\s*(\d+)', text) else None
        fields['Identificador'] = None

        domicilio_match = re.search(r'Domicilio\s*:\s*(.+)', text)
        if domicilio_match:
            fields['Domicilio'] = domicilio_match.group(1).strip()
        else:
            domicilio_lines = [line for line in text.splitlines() if 'Domicilio' in line]
            fields['Domicilio'] = domicilio_lines[0].split(':', 1)[-1].strip() if domicilio_lines else None

        fields['Matrícula'] = None
        fields['Monto total'] = re.search(r'Importe \$\s*:\s*([\d.,]+)', text).group(1) if re.search(r'Importe \$\s*:\s*([\d.,]+)', text) else None
        fields['Tipo de Persona'] = 'Jurídica' if fields['CUIT'] and fields['CUIT'].startswith('3') else 'Física'

        # Additional fields
        fields['Juzgado Nro.'] = re.search(r'Juzgado Nro\.\s*:\s*(\d+)', text).group(1) if re.search(r'Juzgado Nro\.\s*:\s*(\d+)', text) else None
        fields['Juez'] = re.search(r'Juez\s*:\s*(.+)', text).group(1).strip() if re.search(r'Juez\s*:\s*(.+)', text) else None
        fields['Secretaria'] = re.search(r'Secretaria:\s*(.+)', text).group(1).strip() if re.search(r'Secretaria:\s*(.+)', text) else None
        fields['Causa Nro.'] = re.search(r'Causa Nro\.\s*:\s*(\d+)', text).group(1) if re.search(r'Causa Nro\.\s*:\s*(\d+)', text) else None
        fields['Fecha'] = re.search(r'Fecha\s*:\s*(\d{2}/\d{2}/\d{4})', text).group(1) if re.search(r'Fecha\s*:\s*(\d{2}/\d{2}/\d{4})', text) else None
        fields['Fecha de Última Notificación'] = re.search(r'Fecha de Ultima Notificación\s*:\s*(\d{2}/\d{2}/\d{4})', text).group(1) if re.search(r'Fecha de Ultima Notificación\s*:\s*(\d{2}/\d{2}/\d{4})', text) else None

    except Exception as e:
        print(f"Error al extraer datos de {filename}: {e}")
    return fields

def read_pdfs_in_folder(folder_path, output_excel_path):
    """Processes all PDF files in the specified folder and saves the data to an Excel file."""
    data = []

    for filename in os.listdir(folder_path):
        if filename.endswith('.pdf'):
            pdf_path = os.path.join(folder_path, filename)
            print(f'Leyendo contenido de {filename} usando OCR...')
            text = read_pdf_with_ocr(pdf_path)
            fields = extract_fields_from_text(text, filename)
            if fields:
                print(f"Datos extraídos de {filename}: {fields}")
                data.append(fields)

    # Define columns in the desired order
    columns_order = [
        'Archivo', 'Nombre o Razón Social', 'CUIT', 'Repartición', 'Orden',
        'Año', 'Identificador', 'Domicilio', 'Matrícula', 'Monto total',
        'Tipo de Persona', 'Juzgado Nro.', 'Juez', 'Secretaria',
        'Causa Nro.', 'Fecha', 'Fecha de Última Notificación'
    ]

    # Create DataFrame and ensure all columns are included
    df = pd.DataFrame(data)
    df = df.reindex(columns=columns_order)

    # Save to Excel with formatting
    df.to_excel(output_excel_path, index=False)

    workbook = load_workbook(output_excel_path)
    worksheet = workbook.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output_excel_path)
    print(f'Datos guardados exitosamente en {output_excel_path}')

# Ejemplo de uso
folder_path = '/home/peluche/Escritorio/Ramiro Bot/fonseca-bot/titles/Títulos adjuntos causas 2021/'
output_excel_path = 'nuevos_datos_extraidos_completos.xlsx'
read_pdfs_in_folder(folder_path, output_excel_path)
